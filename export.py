"""
Export dữ liệu căn hộ và bài đăng ra file báo cáo Excel.

Usage:
  python export.py             # Xuất báo cáo Excel
"""
import sys
import io
import json
import argparse
import os

sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.stderr.reconfigure(encoding='utf-8', errors='replace')

from config import EXCEL_PATH
from db import ApartmentDB, PostDB


def export_excel():
    """Export dữ liệu ra file Excel với 2 sheet."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("Cần cài openpyxl: pip install openpyxl")
        return

    apt_db = ApartmentDB()
    post_db = PostDB()

    wb = Workbook()

    # === Sheet 1: Tổng hợp chung cư ===
    ws1 = wb.active
    ws1.title = "Chung cư"

    headers1 = ["#", "Tên Dự Án", "Quận", "Địa chỉ", "Km Q1", "Giá tham khảo",
                 "Năm", "Ban công", "Số tin đăng", "Google Maps"]
    _write_header(ws1, headers1)

    for i, apt in enumerate(apt_db.list(), 1):
        post_count = len(post_db.list(apartment_id=apt["id"]))
        balcony = "Có" if apt.get("balcony") is True else ("Tùy căn" if apt.get("balcony") == "tuy_can" else "Không")
        ws1.append([
            i,
            apt.get("name", ""),
            apt.get("district", ""),
            apt.get("address", ""),
            apt.get("km_q1", ""),
            apt.get("price_range", ""),
            apt.get("year", ""),
            balcony,
            post_count,
            apt.get("google_maps", ""),
        ])

    _auto_width(ws1)

    # === Sheet 2: Chi tiết bài đăng ===
    ws2 = wb.create_sheet("Bài đăng")

    headers2 = ["#", "Chung cư", "Quận", "Tiêu đề", "Giá", "Diện tích",
                 "Ngày đăng", "Người đăng", "Nguồn", "Link"]
    _write_header(ws2, headers2)

    for i, post in enumerate(post_db.data, 1):
        apt = apt_db.get(post.get("apartment_id", ""))
        ws2.append([
            i,
            apt["name"] if apt else post.get("apartment_id", ""),
            apt.get("district", "") if apt else "",
            post.get("title", ""),
            post.get("price", ""),
            post.get("area", ""),
            post.get("date", ""),
            post.get("author", ""),
            post.get("source", ""),
            post.get("link", ""),
        ])

    _auto_width(ws2)

    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    wb.save(EXCEL_PATH)
    print(f"Đã export Excel → {EXCEL_PATH}")


def _write_header(ws, headers):
    """Ghi header row với style."""
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _auto_width(ws):
    """Tự động chỉnh độ rộng cột."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def main():
    export_excel()


if __name__ == "__main__":
    main()
